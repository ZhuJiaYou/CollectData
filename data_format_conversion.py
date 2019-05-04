from openpyxl import load_workbook
import re


def xlsx2txt(infilename, outfilename):
    wb = load_workbook(filename=infilename, read_only=True)

    print(type(wb.get_sheet_names()))
    sheet1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    print(sheet1.title)

    sheet = wb.active
    print("ROW is {0}, COL is {1}".format(sheet.max_row, sheet.max_column))

    outfile = open(outfilename, 'a')

    for row in sheet.rows:
        line = ""
        for cell in row:
            newstr = re.sub('[\s]+', ' ', str(cell.value))
            line = line + newstr + "\t"
        line = line.strip()
        line = line + "\n"
        outfile.write(line)

    outfile.close()


if __name__ == '__main__':
#    project = "AspectJ"
#    project = "Birt"
#    project = "Eclipse_Platform_UI"
#    project = "JDT"
#    project = "SWT"
    project = "Tomcat"
    path = "./../datasets/"
    infilename = path + project + ".xlsx"
    outfilename = path + project + ".txt"
    xlsx2txt(infilename, outfilename)
