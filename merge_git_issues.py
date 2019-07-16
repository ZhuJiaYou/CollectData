from openpyxl import load_workbook 
from repository import Repository


with open("/srv/bug_repo_info/selected/selected_repos.txt", "r") as f:
    for line in f:
        repo = Repository(line.strip())
        filename = "{}_{}_{}.txt".format(repo.language, repo.owner, repo.name)
        filename2 = "{}_{}_{}.xlsx".format(repo.language, repo.owner, repo.name)
        with open("/srv/bug_repo_info/selected/merged/{}".format(filename), "w") as fw:
            fw.write("COMMIT_ID\tAUTHOR\tTIME\tISSUE_NUMBER\tI_AUTHOR\tI_TIME\tTITLE\tDESCRIPTION\n")
        wb = load_workbook(filename="/srv/bug_repo_info/selected/is/{}".format(filename2), read_only=True)
        ws = wb.active
        with open("/srv/bug_repo_info/selected/linked/{}".format(filename), "r") as f1:
            for line1 in f1:
                if line1[:9] == "COMMIT_ID":
                    continue
                elem_list = line1.strip().split("\t")
                commit_id = elem_list[0]
                author = elem_list[2]
                time = elem_list[1]
                issue_number = elem_list[3]
                if issue_number != "0":
                # try: 
                    i_author = ws["B"+issue_number].value
                # except AttributeError:
                    # print(line)
                    # print(line1)
                    i_time = ws["C"+issue_number].value
                    title = ws["D"+issue_number].value
                    description = ws["E"+issue_number].value

                    with open("/srv/bug_repo_info/selected/merged/{}".format(filename), "a") as fi:
                        fi.write("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(
                            commit_id, author, time, issue_number, i_author, i_time, title, description))
        print("{} IS FINE.".format(repo.name))

print("ALL IS FINE!")
