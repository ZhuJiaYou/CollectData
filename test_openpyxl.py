from openpyxl import Workbook


wb = Workbook()
ws = wb.active

ws["A1"] = "Hello"
ws["B1"] = "World"
ws["A2"] = "Python"

wb.save('/home/zjy/nlp/CollectData/py/test.xlsx')
