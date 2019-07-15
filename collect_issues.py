from xlsxwriter import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import requests, time, re
from repository import Repository


def run_query(query, headers):
    request = requests.post('https://api.github.com/graphql', json={'query': query}, headers=headers)
    if request.status_code == 200:
        return request.json()
    while request.status_code == 403 or request.status_code == 502:
    # while request.status_code != 200:
        print("WRONG STATUS CODE - {}".format(request.status_code))
        time.sleep(5)
        request = requests.post('https://api.github.com/graphql', json={'query': query}, headers=headers)
    
    if request.status_code != 200:
        raise Exception("Query failed to run by returning code of {0}. {1}".format(request.status_code,
                                                                                   query))
    return request.json()


if __name__ == '__main__':
    REPOS_DIR_PRE = "/srv/bug_repos"

    with open("/srv/bug_repo_info/selected/selected_repos.txt", "r") as f:
        for line in f:
            if line[:6] != "NOTICE" and line[:8] != "LANGUAGE":
                repo = Repository(line.strip())
            else:
                print(line)
                continue

            filename = "{}_{}_{}.xlsx".format(repo.language, repo.owner, repo.name)
            headers = {"Authorization": "token 7810c01eb87dbbcdb989b0d16ff29b3ea4527a21"}
            end_cursor = "null"
            has_next_page = True
            flag = True
            
            wb = Workbook("/srv/bug_repo_info/selected/is/{}".format(filename))
            ws = wb.add_worksheet("issues")

            while has_next_page:
                query = """
                        query {{repository(owner: "{0}", name: "{1}") {{
                          name
                          issues(after: {2}, first: 100) {{
                            nodes {{
                              number
                              title
                              bodyText
                              createdAt
                              author {{
                                login
                              }}
                            }}
                            pageInfo {{
                              endCursor
                              hasNextPage
                            }}
                          }}
                        }}}}
                        """.format(repo.owner, repo.name, end_cursor if flag else '"' + end_cursor + '"')
                flag = False
                result = run_query(query, headers)
                if list(result.keys())[0] != "data":
                    print(repo.language + " - " + repo.owner + " - " + repo.name + " - " + str(result))
                    continue
                print(len(result["data"]["repository"]["issues"]["nodes"]))
                for node in result["data"]["repository"]["issues"]["nodes"]:
                    number = node["number"]
                    title = node["title"]
                    if not node["author"]:
                        author = "ghost"
                    else:
                        author = node["author"]["login"]
                    description = node["bodyText"]
                    create_time = node["createdAt"]
                    # merge_commit = edge["node"]["mergeCommit"]["abbreviatedOid"]

                    title = re.sub(r"\s+", " ", title).strip()
                    author = re.sub(r"\s+", " ", author).strip()
                    description = re.sub(r"\s+", " ", description).strip()
                    data = []
                    data.append(number)
                    data.append(author)
                    data.append(create_time)
                    data.append(title)
                    data.append(description)
                   
                    ws.write_row("A"+str(number), data)

                    """
                    if number == 46935:
                        wb = Workbook("/srv/bug_repo_info/selected/is/{}".format(filename))
                        ws = wb.add_worksheet("issues")
                        ws.write_row("A"+str(number), data)
                        wb.close()

                    if number == 46935:
                        with open("/home/zjy/nlp/CollectData/py/t.txt", "w") as ft:
                            ft.write(title)
                    try:
                        ws["A"+str(number)] = number
                    except IllegalCharacterError:
                        print("*****************************************")
                        print(number)
                        print("*****************************************")
                    try:
                        ws["B"+str(number)] = author
                    except IllegalCharacterError:
                        print("*****************************************")
                        print(author)
                        print("*****************************************")
                    try:
                        ws["C"+str(number)] = create_time
                    except IllegalCharacterError:
                        print("*****************************************")
                        print(create_time)
                        print("*****************************************")
                    try:
                        ws["D"+str(number)] = title
                    except IllegalCharacterError:
                        print("*****************************************")
                        print(number)
                        print(title)
                        print(end_cursor)
                        print("*****************************************")
                        ws["D"+str(number)] = title
                    try:
                        ws["E"+str(number)] = description
                    except IllegalCharacterError:
                        print("*****************************************")
                        print(description)
                        print("*****************************************")
                    """

                end_cursor = result["data"]["repository"]["issues"]["pageInfo"]["endCursor"]
                has_next_page = result["data"]["repository"]["issues"]["pageInfo"]["hasNextPage"]
                print(repo.name + " - " + str(has_next_page))
            # wb.save("/srv/bug_repo_info/selected/is/{}".format(filename))    
            wb.close()
